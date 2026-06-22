

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from models import SinhVien, MonHoc, LopHocPhan, BangDiem
from utils import validate_mssv, validate_ho_ten, validate_ngay_sinh, validate_email


# ============================================================
# CÁC CỘT BẮT BUỘC CHO TỪNG LOẠI FILE EXCEL
# ============================================================
REQUIRED_COLUMNS = {
    "sinh_vien": [
        "MSSV", "Họ tên", "Ngày sinh", "Giới tính", "Lớp", "Email"
    ],
    "mon_hoc": [
        "Mã môn", "Tên môn", "Số tín chỉ"
    ],
    "lop_hoc_phan": [
        "Mã LHP", "Mã môn", "Học kỳ", "Năm học", "Danh sách MSSV (phân cách bởi dấu phẩy)"
    ],
    "diem": [
        "MSSV", "Mã môn", "Điểm quá trình", "Điểm thi"
    ],
}


def validate_excel_columns(ws, loai_file):

    header_row = []
    for cell in ws[1]:
        val = cell.value
        if val is not None:
            header_row.append(str(val).strip())
        else:
            header_row.append("")

    non_empty_headers = [h for h in header_row if h]
    if len(non_empty_headers) < len(required):
        return False, (
            f"File Excel thiếu cột! File cần có {len(required)} cột:\n"
            + ", ".join(f"'{c}'" for c in required)
            + f"\n\nFile của bạn chỉ có {len(non_empty_headers)} cột."
        )

    header_lower = [h.lower() for h in header_row]
    missing_cols = []
    for req_col in required:
        if req_col.lower() not in header_lower:
            missing_cols.append(req_col)

    if missing_cols:
        return False, (
            f"File Excel không đúng định dạng!\n\n"
            f"Cột bị thiếu hoặc sai tên:\n"
            + "\n".join(f"  ❌ '{c}'" for c in missing_cols)
            + f"\n\nFile cần có đầy đủ các cột:\n"
            + ", ".join(f"'{c}'" for c in required)
        )

    return True, ""


def import_lop_hoc_phan(filepath):
   
    ds_lhp = []
    errors = []
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        valid, err_msg = validate_excel_columns(ws, "lop_hoc_phan")
        if not valid:
            wb.close()
            errors.append(err_msg)
            return ds_lhp, errors
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or all(cell is None for cell in row):
                continue
            
            try:
                ma_lhp = str(row[0]).strip() if row[0] is not None else ""
                ma_mon = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                hoc_ky = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                nam_hoc = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
                
                # Đọc danh sách MSSV (cách nhau bởi dấu phẩy)
                ds_mssv = []
                if len(row) > 4 and row[4] is not None:
                    raw = str(row[4]).strip()
                    if raw and raw != "None":
                        ds_mssv = [m.strip() for m in raw.split(",") if m.strip()]
                
                if not ma_lhp or ma_lhp == "None":
                    continue
                if not ma_mon or ma_mon == "None":
                    errors.append(f"Hàng {row_idx}: Thiếu mã môn cho LHP '{ma_lhp}'")
                    continue
                
                lhp = LopHocPhan(ma_lhp, ma_mon, hoc_ky, nam_hoc, ds_mssv)
                ds_lhp.append(lhp)
                
            except Exception as e:
                errors.append(f"Hàng {row_idx}: Lỗi đọc dữ liệu - {str(e)}")
        
        wb.close()
        
    except FileNotFoundError:
        errors.append(f"Không tìm thấy file: {filepath}")
    except Exception as e:
        errors.append(f"Lỗi đọc file Excel: {str(e)}")
    
    return ds_lhp, errors


def export_template_lop_hoc_phan(filepath):
   
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách Lớp Học Phần"
    
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    headers = ["Mã LHP", "Mã môn", "Học kỳ", "Năm học", "Danh sách MSSV (phân cách bởi dấu phẩy)"]
    col_widths = [15, 15, 10, 15, 45]
    
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    sample_data = [
        ["LHP001", "CS101", "HK1", "2024-2025", "SV001, SV002, SV003"],
        ["LHP002", "CS102", "HK1", "2024-2025", "SV001, SV004, SV005"],
        ["LHP003", "CS103", "HK2", "2024-2025", "SV002, SV003"],
        ["LHP004", "MATH01", "HK1", "2024-2025", "SV001, SV002, SV004, SV005"],
    ]
    
    data_align = Alignment(horizontal="center", vertical="center")
    
    for row_idx, data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx != 5:  # Không căn giữa cột MSSV dài
                cell.alignment = data_align
    
    wb.save(filepath)
    wb.close()


def import_sinh_vien(filepath):
    
    
    ds_sv = []
    errors = []
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active  # Lấy sheet đầu tiên

        valid, err_msg = validate_excel_columns(ws, "sinh_vien")
        if not valid:
            wb.close()
            errors.append(err_msg)
            return ds_sv, errors
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or all(cell is None for cell in row):
                continue  # Bỏ qua hàng trống
            
            try:
                mssv = str(row[0]).strip() if row[0] is not None else ""
                ho_ten = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                ngay_sinh = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                gioi_tinh = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
                lop = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
                email = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""
                
                missing = []
                if not mssv or mssv == "None":
                    missing.append("MSSV")
                if not ho_ten or ho_ten == "None":
                    missing.append("Họ tên")
                if not ngay_sinh or ngay_sinh == "None":
                    missing.append("Ngày sinh")
                if not gioi_tinh or gioi_tinh == "None":
                    missing.append("Giới tính")
                if not lop or lop == "None":
                    missing.append("Lớp")
                if not email or email == "None":
                    missing.append("Email")
                
                if missing:
                    errors.append(f"Hàng {row_idx}: Thiếu thông tin ({', '.join(missing)})")
                    continue
                
                row_errors = []
                ok_mssv, msg_mssv = validate_mssv(mssv)
                if not ok_mssv:
                    row_errors.append(f"MSSV '{mssv}' không hợp lệ ({msg_mssv})")
                
                ok_hoten, msg_hoten = validate_ho_ten(ho_ten)
                if not ok_hoten:
                    row_errors.append(f"Họ tên '{ho_ten}' không hợp lệ ({msg_hoten})")
                
                ok_ngaysinh, msg_ngaysinh = validate_ngay_sinh(ngay_sinh)
                if not ok_ngaysinh:
                    row_errors.append(f"Ngày sinh '{ngay_sinh}' không hợp lệ ({msg_ngaysinh})")
                
                if gioi_tinh not in ("Nam", "Nữ"):
                    row_errors.append(f"Giới tính '{gioi_tinh}' không hợp lệ (phải là 'Nam' hoặc 'Nữ')")
                
                ok_email, msg_email = validate_email(email, ho_ten, mssv)
                if not ok_email:
                    row_errors.append(f"Email '{email}' không hợp lệ ({msg_email})")
                
                if row_errors:
                    errors.append(f"Hàng {row_idx}: {'; '.join(row_errors)}")
                    continue
                
                sv = SinhVien(mssv, ho_ten, ngay_sinh, gioi_tinh, lop, email)
                ds_sv.append(sv)
                
            except Exception as e:
                errors.append(f"Hàng {row_idx}: Lỗi đọc dữ liệu - {str(e)}")
        
        wb.close()
        
    except FileNotFoundError:
        errors.append(f"Không tìm thấy file: {filepath}")
    except Exception as e:
        errors.append(f"Lỗi đọc file Excel: {str(e)}")
    
    return ds_sv, errors


def import_mon_hoc(filepath):
   
    ds_mh = []
    errors = []
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        valid, err_msg = validate_excel_columns(ws, "mon_hoc")
        if not valid:
            wb.close()
            errors.append(err_msg)
            return ds_mh, errors
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or all(cell is None for cell in row):
                continue
            
            try:
                ma_mon = str(row[0]).strip() if row[0] is not None else ""
                ten_mon = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                so_tc = int(row[2]) if len(row) > 2 and row[2] is not None else 0
                
                if not ma_mon or ma_mon == "None":
                    continue
                if not ten_mon or ten_mon == "None":
                    errors.append(f"Hàng {row_idx}: Thiếu tên môn cho mã '{ma_mon}'")
                    continue
                
                mh = MonHoc(ma_mon, ten_mon, so_tc)
                ds_mh.append(mh)
                
            except Exception as e:
                errors.append(f"Hàng {row_idx}: Lỗi đọc dữ liệu - {str(e)}")
        
        wb.close()
        
    except Exception as e:
        errors.append(f"Lỗi đọc file Excel: {str(e)}")
    
    return ds_mh, errors


def import_diem(filepath):
   
    ds_diem = []
    errors = []
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        valid, err_msg = validate_excel_columns(ws, "diem")
        if not valid:
            wb.close()
            errors.append(err_msg)
            return ds_diem, errors
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or all(cell is None for cell in row):
                continue
            
            try:
                mssv = str(row[0]).strip() if row[0] is not None else ""
                ma_mon = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                diem_qt = float(row[2]) if len(row) > 2 and row[2] is not None else 0.0
                diem_thi = float(row[3]) if len(row) > 3 and row[3] is not None else 0.0
                
                if not mssv or mssv == "None":
                    continue
                if not ma_mon or ma_mon == "None":
                    errors.append(f"Hàng {row_idx}: Thiếu mã môn cho MSSV '{mssv}'")
                    continue
                
                if diem_qt < 0 or diem_qt > 10:
                    errors.append(f"Hàng {row_idx}: Điểm QT không hợp lệ ({diem_qt})")
                    continue
                if diem_thi < 0 or diem_thi > 10:
                    errors.append(f"Hàng {row_idx}: Điểm thi không hợp lệ ({diem_thi})")
                    continue
                
                bd = BangDiem(mssv, ma_mon, round(diem_qt, 2), round(diem_thi, 2))
                ds_diem.append(bd)
                
            except Exception as e:
                errors.append(f"Hàng {row_idx}: Lỗi đọc dữ liệu - {str(e)}")
        
        wb.close()
        
    except Exception as e:
        errors.append(f"Lỗi đọc file Excel: {str(e)}")
    
    return ds_diem, errors


def export_template_sinh_vien(filepath):
   
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách Sinh viên"
    
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    headers = ["MSSV", "Họ tên", "Ngày sinh", "Giới tính", "Lớp", "Email"]
    col_widths = [12, 25, 15, 12, 15, 30]
    
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    sample_data = [
        ["SV001", "Nguyễn Văn An", "15/03/2003", "Nam", "CNTT01", "an.nv@email.com"],
        ["SV002", "Trần Thị Bình", "22/07/2003", "Nữ", "CNTT01", "binh.tt@email.com"],
        ["SV003", "Lê Hoàng Cường", "10/01/2004", "Nam", "CNTT02", "cuong.lh@email.com"],
        ["SV004", "Phạm Minh Duy", "05/11/2003", "Nam", "CNTT02", "duy.pm@email.com"],
        ["SV005", "Hoàng Thị Em", "30/09/2003", "Nữ", "CNTT01", "em.ht@email.com"],
    ]
    
    data_align = Alignment(horizontal="center", vertical="center")
    
    for row_idx, data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx != 2 and col_idx != 6:  # Không căn giữa họ tên và email
                cell.alignment = data_align
    
    wb.save(filepath)
    wb.close()


def export_template_mon_hoc(filepath):
 
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách Môn học"
    
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    headers = ["Mã môn", "Tên môn", "Số tín chỉ"]
    col_widths = [15, 35, 15]
    
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    sample_data = [
        ["CS101", "Nhập môn lập trình", 3],
        ["CS102", "Cấu trúc dữ liệu và giải thuật", 4],
        ["CS103", "Cơ sở dữ liệu", 3],
        ["CS104", "Mạng máy tính", 3],
        ["MATH01", "Toán cao cấp", 4],
    ]
    
    data_align = Alignment(horizontal="center", vertical="center")
    
    for row_idx, data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx != 2:
                cell.alignment = data_align
    
    wb.save(filepath)
    wb.close()


def export_template_diem(filepath):
   
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bảng điểm"
    
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    headers = ["MSSV", "Mã môn", "Điểm quá trình", "Điểm thi"]
    col_widths = [12, 12, 18, 12]
    
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    sample_data = [
        ["SV001", "CS101", 8.5, 7.0],
        ["SV001", "CS102", 7.0, 8.5],
        ["SV002", "CS101", 9.0, 8.0],
        ["SV002", "CS102", 6.5, 7.5],
        ["SV003", "CS101", 8.0, 9.0],
    ]
    
    data_align = Alignment(horizontal="center", vertical="center")
    
    for row_idx, data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = data_align
    
    wb.save(filepath)
    wb.close()


# ============================================================
# XUẤT DỮ LIỆU THỰC TẾ RA FILE EXCEL
# ============================================================

def _apply_header_style(ws, headers, col_widths, fill_color):
   
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    return thin_border


def export_data_sinh_vien(filepath, ds_sinh_vien):
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách Sinh viên"

    headers = ["MSSV", "Họ tên", "Ngày sinh", "Giới tính", "Lớp", "Email"]
    col_widths = [12, 28, 15, 12, 15, 32]
    thin_border = _apply_header_style(ws, headers, col_widths, "4472C4")

    data_align_center = Alignment(horizontal="center", vertical="center")

    for row_idx, sv in enumerate(ds_sinh_vien, start=2):
        row_data = [sv.mssv, sv.ho_ten, sv.ngay_sinh, sv.gioi_tinh, sv.lop, sv.email]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx not in (2, 6):
                cell.alignment = data_align_center

    wb.save(filepath)
    wb.close()
    return len(ds_sinh_vien)


def export_data_mon_hoc(filepath, ds_mon_hoc):
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách Môn học"

    headers = ["Mã môn", "Tên môn", "Số tín chỉ"]
    col_widths = [15, 40, 15]
    thin_border = _apply_header_style(ws, headers, col_widths, "548235")

    data_align_center = Alignment(horizontal="center", vertical="center")

    for row_idx, mh in enumerate(ds_mon_hoc, start=2):
        row_data = [mh.ma_mon, mh.ten_mon, mh.so_tin_chi]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx != 2:  # Căn giữa trừ Tên môn
                cell.alignment = data_align_center

    wb.save(filepath)
    wb.close()
    return len(ds_mon_hoc)


def export_data_diem(filepath, ds_diem, ds_sinh_vien=None, ds_mon_hoc=None):
   
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bảng điểm"

    sv_lookup = {}
    if ds_sinh_vien:
        for sv in ds_sinh_vien:
            sv_lookup[sv.mssv.upper()] = sv.ho_ten

    mh_lookup = {}
    if ds_mon_hoc:
        for mh in ds_mon_hoc:
            mh_lookup[mh.ma_mon.upper()] = (mh.ten_mon, mh.so_tin_chi)

    headers = [
        "MSSV", "Họ tên", "Mã môn", "Tên môn", "Số tín chỉ",
        "Điểm quá trình", "Điểm thi", "Điểm tổng kết", "Hệ 4", "Điểm chữ", "Xếp loại"
    ]
    col_widths = [12, 28, 12, 35, 12, 16, 12, 15, 10, 12, 15]
    thin_border = _apply_header_style(ws, headers, col_widths, "BF8F00")

    data_align_center = Alignment(horizontal="center", vertical="center")

    for row_idx, bd in enumerate(ds_diem, start=2):
        ho_ten = sv_lookup.get(bd.mssv.upper(), "")
        ten_mon, so_tc = mh_lookup.get(bd.ma_mon.upper(), ("", ""))

        row_data = [
            bd.mssv, ho_ten, bd.ma_mon, ten_mon, so_tc,
            bd.diem_qua_trinh, bd.diem_thi,
            bd.diem_tong_ket, bd.diem_he4, bd.diem_chu, bd.xep_loai
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx not in (2, 4):
                cell.alignment = data_align_center

    wb.save(filepath)
    wb.close()
    return len(ds_diem)
